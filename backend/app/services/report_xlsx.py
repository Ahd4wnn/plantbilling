"""Styled Excel (.xlsx) builder for the manager sales report.

Managers said the CSV "cannot even be called a report", so this produces a
polished, multi-tab workbook: a Summary landing tab plus one tab each for
bills, line items, customers, staff, expenses and the edit/delete log.

Design goals (kept deliberately professional, brand-consistent):
- Brand-orange header bands (#F05B01) with white bold headers, matching the app.
- Money stored as REAL NUMBERS with an Indian ₹ format ("₹#,##,##0.00") so Excel
  can sum/sort/filter — no "INR 123.00" text like the old CSV, and no mojibake
  (xlsx stores the ₹ in XML, unlike CSV which needs a BOM and still guesses).
- Frozen header rows, auto-filters, sensible auto-fit column widths, zebra
  banding and thin borders so a manager can read or print it as-is.

The caller (routers/bills.py) assembles the data; this module only formats it.
"""
from __future__ import annotations

import calendar
import datetime as dt
import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Any, Callable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Brand palette ───────────────────────────────────────────────────────────
# Must track Brand.Orange in android/.../ui/theme/Color.kt. The trailing comments
# carry the previous palettes, to paste back if the brand ever reverts.
BRAND = "F05B01"          # brand orange (matches the app + launcher icon)  [green: 2E6F40]
BRAND_DARK = "C24700"     # deeper orange for the title band                [green: 1E4D2B]
BAND = "FEF0E6"           # pale orange zebra stripe                        [green: EAF3EC]
GRID = "F2DECF"           # soft grid line                                  [green: D8E2DA]
TEXT_DARK = "2E1A12"      # near-black warm-tinted body text                [green: 1A2E22]
MUTED = "6B5A50"          # secondary labels                                [green: 5B6B60]

# Attendance register cell fills (§ build_attendance_register). Deliberately
# outside the brand: red always means "absent", whatever colour the app is.
ABSENT_FILL = "FDE0DE"    # light red behind an absence
ABSENT_TEXT = "9F1239"    # dark red so the A itself is legible, not just the cell
HALF_FILL = "FEF0D0"      # light amber behind a half-day
HALF_TEXT = "8A5A00"
PRESENT_TEXT = "1E6B3C"   # calm green tick-equivalent for a P
UNMARKED_TEXT = "B0A79F"  # a day nobody marked is not an absence — keep it faint

INR_FMT = '₹#,##,##0.00'   # ₹ with Indian lakh grouping, 2dp
INT_FMT = '#,##0'

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
_SUB_FONT = Font(name="Calibri", size=10, color="FFFFFF")
_BODY_FONT = Font(name="Calibri", size=11, color=TEXT_DARK)
_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=TEXT_DARK)
_MUTED_FONT = Font(name="Calibri", size=10, color=MUTED)
_KPI_FONT = Font(name="Calibri", size=13, bold=True, color=BRAND_DARK)

_HEADER_FILL = PatternFill("solid", fgColor=BRAND)
_TITLE_FILL = PatternFill("solid", fgColor=BRAND_DARK)
_BAND_FILL = PatternFill("solid", fgColor=BAND)
_KPI_FILL = PatternFill("solid", fgColor="FFF8F3")

_thin = Side(style="thin", color=GRID)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")


# Column spec: (header, kind, width_hint)
#   kind: "text" | "money" | "int" | "wraptext"
Col = tuple[str, str, int]


def _num(v: Any) -> float:
    if isinstance(v, Decimal):
        return float(v)
    return float(v or 0)


def _autofit(ws: Worksheet, widths: dict[int, int]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_table(
    ws: Worksheet,
    top: int,
    title: str,
    cols: Sequence[Col],
    rows: Sequence[Sequence[Any]],
    empty_text: str = "(nothing to show for this period)",
) -> int:
    """Render a titled, styled table starting at row `top`. Returns next free row."""
    ncols = len(cols)
    # Section title band
    ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=ncols)
    tcell = ws.cell(row=top, column=1, value=title)
    tcell.font = Font(name="Calibri", size=12, bold=True, color=BRAND_DARK)
    tcell.alignment = _LEFT
    ws.row_dimensions[top].height = 22

    header_row = top + 1
    for c, (label, _kind, _w) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[header_row].height = 20

    r = header_row + 1
    if not rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        ec = ws.cell(row=r, column=1, value=empty_text)
        ec.font = _MUTED_FONT
        ec.alignment = _LEFT
        ec.border = _BORDER
        return r + 2

    for i, row in enumerate(rows):
        band = _BAND_FILL if i % 2 else None
        for c, ((_label, kind, _w), value) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=r, column=c)
            if kind == "money":
                cell.value = _num(value)
                cell.number_format = INR_FMT
                cell.alignment = _RIGHT
            elif kind == "int":
                cell.value = int(value or 0)
                cell.number_format = INT_FMT
                cell.alignment = _RIGHT
            elif kind == "wraptext":
                cell.value = value
                cell.alignment = _LEFT_WRAP
            else:
                cell.value = value
                cell.alignment = _LEFT
            cell.font = _BODY_FONT
            cell.border = _BORDER
            if band:
                cell.fill = band
        r += 1

    # Freeze header + auto-filter the data block
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"{get_column_letter(1)}{header_row}:{get_column_letter(ncols)}{r - 1}"
    )
    _autofit(ws, {i: w for i, (_l, _k, w) in enumerate(cols, start=1)})
    return r + 2


def _build_summary_tab(ws: Worksheet, meta: dict, kpis: list[tuple[str, Any, str]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 3

    # Title band (B2:C4)
    ws.merge_cells("B2:C2")
    t = ws.cell(row=2, column=2, value="PLANTBILL")
    t.font = Font(name="Calibri", size=10, bold=True, color="FFD2B3")
    t.alignment = _LEFT
    ws.merge_cells("B3:C3")
    t2 = ws.cell(row=3, column=2, value="Sales Report")
    t2.font = _TITLE_FONT
    t2.alignment = _LEFT
    for rr in (2, 3, 4):
        for cc in (2, 3):
            ws.cell(row=rr, column=cc).fill = _TITLE_FILL
    ws.merge_cells("B4:C4")
    t3 = ws.cell(row=4, column=2, value=meta["shop_name"])
    t3.font = _SUB_FONT
    t3.alignment = _LEFT
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 18

    # Meta block
    meta_rows = [
        ("Period", meta["period"]),
        ("Generated by", meta["generated_by"]),
        ("Generated at", meta["generated_at"]),
    ]
    if meta.get("staff_filter"):
        meta_rows.append(("Filtered to staff", meta["staff_filter"]))
    r = 6
    for label, value in meta_rows:
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _MUTED_FONT
        lc.alignment = _LEFT
        vc = ws.cell(row=r, column=3, value=value)
        vc.font = _LABEL_FONT
        vc.alignment = _LEFT
        r += 1

    # KPI cards (label / value pairs)
    r += 1
    hdr = ws.cell(row=r, column=2, value="AT A GLANCE")
    hdr.font = Font(name="Calibri", size=11, bold=True, color=BRAND)
    r += 1
    for label, value, kind in kpis:
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _LABEL_FONT
        lc.alignment = _LEFT
        lc.fill = _KPI_FILL
        lc.border = _BORDER
        vc = ws.cell(row=r, column=3)
        if kind == "money":
            vc.value = _num(value)
            vc.number_format = INR_FMT
        elif kind == "int":
            vc.value = int(value or 0)
            vc.number_format = INT_FMT
        else:
            vc.value = value
        vc.font = _KPI_FONT
        vc.alignment = _RIGHT
        vc.fill = _KPI_FILL
        vc.border = _BORDER
        ws.row_dimensions[r].height = 20
        r += 1


@dataclass(frozen=True)
class AttendanceMark:
    """One worker's status on one day, as stored. The register is pivoted from these."""

    day: dt.date
    worker: str
    joined_on: dt.date | None
    status: str  # present | absent | half_day


# What each status looks like in a cell: (letter, fill, text colour, bold).
_MARK_STYLE = {
    "present": ("P", None, PRESENT_TEXT, False),
    "half_day": ("H", HALF_FILL, HALF_TEXT, True),
    "absent": ("A", ABSENT_FILL, ABSENT_TEXT, True),
}
# A day nobody marked. Not an absence — a worker isn't docked for a day the
# manager forgot to open the app, so it must not look like one.
_UNMARKED = ("·", None, UNMARKED_TEXT, False)
# Before the worker joined the shop. Blanked out, not dotted — there was nothing
# to mark, so it must not read as a gap in the manager's record-keeping.
_PRE_JOINING = ("", "F2F0ED", UNMARKED_TEXT, False)

_DAY_COL_WIDTH = 3.4
_NAME_COL_WIDTH = 24
_JOINED_COL_WIDTH = 13
_TOTAL_COL_WIDTH = 7


def _months_between(start: dt.date, end: dt.date) -> list[tuple[int, int]]:
    """Every (year, month) the range touches, in order."""
    months: list[tuple[int, int]] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append((y, m))
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return months


def _build_attendance_register(
    ws: Worksheet,
    marks: Sequence[AttendanceMark],
    period: tuple[dt.date, dt.date] | None,
    roster: Sequence[tuple[str, dt.date | None]] = (),
) -> None:
    """A school-style attendance register: workers down the side, days across the
    top, one block per calendar month.

    The old tab was a row per record, so answering "was Ramesh in on the 12th?"
    meant scanning hundreds of rows. This is the shape a manager already knows how
    to read — and can print and hang up.
    """
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    if not marks and not roster:
        _write_table(
            ws, 1, "ATTENDANCE REGISTER",
            [("Attendance", "text", 60)], [],
            "(no attendance marked in this period)",
        )
        return

    by_worker_day = {(m.worker, m.day): m.status for m in marks}
    # The roster is the register's enrolment list, like a school's. Anyone with a
    # mark but no roster row (a worker removed since) is kept so their history
    # doesn't vanish from an already-issued report.
    joined_by_worker: dict[str, dt.date | None] = {m.worker: m.joined_on for m in marks}
    joined_by_worker.update({name: joined for name, joined in roster})
    workers = sorted(joined_by_worker)

    days_present = [m.day for m in marks]
    start = period[0] if period else min(days_present, default=dt.date.today())
    end = period[1] if period else max(days_present, default=dt.date.today())

    row = 1
    grand_absent = 0
    first_header: int | None = None
    blocks = 0
    for year, month in _months_between(start, end):
        month_start = max(start, dt.date(year, month, 1))
        last_day = calendar.monthrange(year, month)[1]
        month_end = min(end, dt.date(year, month, last_day))
        days = [
            month_start + dt.timedelta(days=i)
            for i in range((month_end - month_start).days + 1)
        ]
        # Workers who hadn't joined yet by the end of this month don't belong in
        # its block at all.
        in_month = [
            w for w in workers
            if (joined_by_worker.get(w) or month_start) <= month_end
            or any((w, d) in by_worker_day for d in days)
        ]
        if not in_month:
            continue

        if first_header is None:
            first_header = row + 1  # the block's header row
        blocks += 1
        row = _write_register_block(ws, row, year, month, days, in_month, joined_by_worker, by_worker_day)
        grand_absent += sum(
            1 for w in workers for d in days if by_worker_day.get((w, d)) == "absent"
        )

    # A sheet has exactly one frozen pane, so with several month blocks only the
    # name columns can stay put — freezing one block's header would strand the
    # others. A single month gets the full treatment, header row included.
    if blocks == 1 and first_header is not None:
        ws.freeze_panes = ws.cell(row=first_header + 1, column=3)
        ws.print_title_rows = f"{first_header}:{first_header}"
    else:
        ws.freeze_panes = "C1"

    # The one number a manager is actually looking for.
    total = ws.cell(row=row, column=1, value="TOTAL ABSENT DAYS (all workers, whole period)")
    total.font = Font(name="Calibri", size=12, bold=True, color=ABSENT_TEXT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    value = ws.cell(row=row, column=5, value=grand_absent)
    value.font = Font(name="Calibri", size=12, bold=True, color=ABSENT_TEXT)
    value.fill = PatternFill("solid", fgColor=ABSENT_FILL)
    value.alignment = _CENTER
    value.border = _BORDER
    ws.row_dimensions[row].height = 22


def _write_register_block(
    ws: Worksheet,
    top: int,
    year: int,
    month: int,
    days: list[dt.date],
    workers: list[str],
    joined_by_worker: dict[str, dt.date | None],
    by_worker_day: dict[tuple[str, dt.date], str],
) -> int:
    """One month's grid. Returns the next free row."""
    first_day_col = 3
    ncols = first_day_col + len(days) + 4  # + P / H / A / Days

    # Month heading
    ws.merge_cells(start_row=top, start_column=1, end_row=top, end_column=ncols)
    title = ws.cell(row=top, column=1, value=f"{calendar.month_name[month].upper()} {year}")
    title.font = Font(name="Calibri", size=13, bold=True, color=BRAND_DARK)
    title.alignment = _LEFT
    ws.row_dimensions[top].height = 24

    header = top + 1
    labels = ["Worker", "Joined"] + [str(d.day) for d in days] + ["P", "H", "A", "Days"]
    for c, label in enumerate(labels, start=1):
        cell = ws.cell(row=header, column=c, value=label)
        cell.font = _HEADER_FONT
        # Weekend columns get a darker band, as in a paper register.
        weekend = first_day_col <= c < first_day_col + len(days) and days[c - first_day_col].weekday() >= 5
        cell.fill = PatternFill("solid", fgColor=BRAND_DARK) if weekend else _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[header].height = 20

    r = header + 1
    for i, worker in enumerate(workers):
        band = _BAND_FILL if i % 2 else None
        name_cell = ws.cell(row=r, column=1, value=worker)
        name_cell.font = _LABEL_FONT
        name_cell.alignment = _LEFT
        name_cell.border = _BORDER
        if band:
            name_cell.fill = band

        joined = joined_by_worker.get(worker)
        joined_cell = ws.cell(row=r, column=2, value=joined.strftime("%d %b %y") if joined else "—")
        joined_cell.font = _MUTED_FONT
        joined_cell.alignment = _CENTER
        joined_cell.border = _BORDER
        if band:
            joined_cell.fill = band

        present = half = absent = 0
        for j, day in enumerate(days):
            status = by_worker_day.get((worker, day))
            if status is None and joined is not None and day < joined:
                letter, fill, colour, bold = _PRE_JOINING
            else:
                letter, fill, colour, bold = _MARK_STYLE.get(status, _UNMARKED)
            if status == "present":
                present += 1
            elif status == "half_day":
                half += 1
            elif status == "absent":
                absent += 1

            cell = ws.cell(row=r, column=first_day_col + j, value=letter)
            cell.font = Font(name="Calibri", size=11, bold=bold, color=colour)
            cell.alignment = _CENTER
            cell.border = _BORDER
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            elif band:
                cell.fill = band

        # Days worked matches the app's ledger: present + ½·half-day.
        worked = present + half * 0.5
        totals = [
            (present, PRESENT_TEXT), (half, HALF_TEXT),
            (absent, ABSENT_TEXT), (worked, TEXT_DARK),
        ]
        for k, (value, colour) in enumerate(totals):
            cell = ws.cell(row=r, column=first_day_col + len(days) + k, value=value)
            cell.font = Font(name="Calibri", size=11, bold=True, color=colour)
            cell.number_format = "0.#"
            cell.alignment = _CENTER
            cell.border = _BORDER
            if k == 2 and absent:
                cell.fill = PatternFill("solid", fgColor=ABSENT_FILL)
            elif band:
                cell.fill = band
        r += 1

    # Footer: how many workers were absent on each day.
    foot = ws.cell(row=r, column=1, value="Absent that day")
    foot.font = Font(name="Calibri", size=10, bold=True, color=ABSENT_TEXT)
    foot.alignment = _LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for j, day in enumerate(days):
        count = sum(1 for w in workers if by_worker_day.get((w, day)) == "absent")
        cell = ws.cell(row=r, column=first_day_col + j, value=count or None)
        cell.font = Font(name="Calibri", size=10, bold=True, color=ABSENT_TEXT)
        cell.alignment = _CENTER
        cell.border = _BORDER
    month_absent = sum(
        1 for w in workers for d in days if by_worker_day.get((w, d)) == "absent"
    )
    tot = ws.cell(row=r, column=first_day_col + len(days) + 2, value=month_absent)
    tot.font = Font(name="Calibri", size=11, bold=True, color=ABSENT_TEXT)
    tot.fill = PatternFill("solid", fgColor=ABSENT_FILL)
    tot.alignment = _CENTER
    tot.border = _BORDER
    r += 1

    legend = ws.cell(
        row=r, column=1,
        value=(
            "P = Present    H = Half day    A = Absent (shaded red)    "
            "· = Not marked    (grey) = before they joined"
        ),
    )
    legend.font = _MUTED_FONT
    legend.alignment = _LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(ncols, 14))
    r += 2

    widths = {1: _NAME_COL_WIDTH, 2: _JOINED_COL_WIDTH}
    widths.update({first_day_col + j: _DAY_COL_WIDTH for j in range(len(days))})
    widths.update({first_day_col + len(days) + k: _TOTAL_COL_WIDTH for k in range(4)})
    _autofit(ws, widths)
    return r


def build_report_workbook(
    *,
    meta: dict,
    kpis: list[tuple[str, Any, str]],
    bills: Sequence[Sequence[Any]],
    line_items: Sequence[Sequence[Any]],
    customers: Sequence[Sequence[Any]],
    staff: Sequence[Sequence[Any]],
    expenses: Sequence[Sequence[Any]],
    expenses_by_category: Sequence[Sequence[Any]] = (),
    labour: Sequence[Sequence[Any]] = (),
    attendance: Sequence[AttendanceMark] = (),
    attendance_range: tuple[dt.date, dt.date] | None = None,
    attendance_roster: Sequence[tuple[str, dt.date | None]] = (),
    audit: Sequence[Sequence[Any]] = (),
) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.sheet_properties.tabColor = BRAND
    _build_summary_tab(summary, meta, kpis)

    tabs = [
        ("All Bills", [
            ("Bill No", "text", 12), ("Date & Time", "text", 18), ("Staff", "text", 22),
            ("Customer", "text", 20), ("Phone", "text", 15), ("Items", "wraptext", 40),
            ("Subtotal", "money", 14), ("Discount", "money", 13), ("Total", "money", 14),
            ("Cash", "money", 13), ("UPI", "money", 13), ("Due", "money", 13),
            ("Payment", "text", 11), ("Edited", "text", 9),
        ], bills, "(no bills in this period)"),
        ("Line Items", [
            ("Bill No", "text", 12), ("Date & Time", "text", 18), ("Product", "text", 30),
            ("Unit Price", "money", 14), ("Qty", "int", 8), ("Line Total", "money", 15),
        ], line_items, "(no line items in this period)"),
        ("Customers", [
            ("Customer", "text", 26), ("Phone", "text", 16), ("Bills", "int", 9),
            ("Total Purchased", "money", 18), ("Total Due", "money", 15),
            ("Staff Handled", "wraptext", 30),
        ], customers, "(no bills in this period)"),
        ("Staff", [
            ("Staff", "text", 28), ("Bills", "int", 9), ("Total Sales", "money", 16),
            ("Cash", "money", 14), ("UPI", "money", 14), ("Due", "money", 14),
        ], staff, "(no bills in this period)"),
        ("Expenses by Category", [
            ("Category", "text", 30), ("Count", "int", 9), ("Total", "money", 16),
        ], expenses_by_category, "(no expenses in this period)"),
        ("Expenses", [
            ("Date & Time", "text", 18), ("Recorded By", "text", 26),
            ("Category", "text", 26), ("Remark", "wraptext", 34), ("Amount", "money", 15),
        ], expenses, "(no expenses in this period)"),
        ("Labour Payments", [
            ("Date & Time", "text", 18), ("Labourer", "text", 22), ("Gender", "text", 9),
            ("Type", "text", 11), ("Days", "text", 7), ("Wage", "money", 12), ("Total", "money", 13),
            ("Method", "text", 10), ("Cash", "money", 12), ("UPI", "money", 12),
            ("Due", "money", 12), ("Recorded By", "text", 24), ("Note", "wraptext", 28),
        ], labour, "(no labour payments in this period)"),
        ("Edit & Delete Log", [
            ("Date & Time", "text", 18), ("Action", "text", 10), ("Bill No", "text", 12),
            ("By", "text", 26), ("Details", "wraptext", 50),
        ], audit, "(no edits or deletions in this period)"),
    ]

    for name, cols, rows, empty in tabs:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        _write_table(ws, 1, name.upper(), cols, rows, empty)

    # The register goes just before the audit log so attendance sits next to the
    # labour payments it explains.
    register = wb.create_sheet("Attendance Register", wb.sheetnames.index("Edit & Delete Log"))
    _build_attendance_register(register, attendance, attendance_range, attendance_roster)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
